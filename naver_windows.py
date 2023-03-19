import re, time, os, datetime
import threading
from tkinter import *
from tkinter import filedialog
import openpyxl
import tkinter.messagebox
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException

master = Tk()
master.minsize(width=570, height=101)
master.columnconfigure(1, weight=1)

#os.chdir('C:\\Git\\bonitosan_automates')
browser = webdriver.Chrome('driver\chromedriver')
delay = 5

# ======================================================================================================== #
#                                                Webdriver Functions                                       #
# ======================================================================================================== #

def logme(text):
    logfile = open('naver_runtime.log', 'a+')
    x = datetime.datetime.now()
    logfile.write(str(x) + '\t' + text + '\n')
    logfile.close()

def write_lnk(text):
    logfile = open('output.txt', 'a+')
    x = datetime.datetime.now()
    logfile.write(str(x) + '\t' + text + '\n')
    logfile.close()

def write(input,output_file):
    encode_input = input.encode('ascii', 'ignore')
    with open(output_file, 'wb') as f:
        f.write(encode_input)
        f.close()

def chk_exist(xpath):
    try:
        browser.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True

def updstat(current_status):
    status.configure(text=current_status)

def loadact(xlsx):
    loadfile = openpyxl.load_workbook(xlsx)

    # *** check header *** #
    actlist = list(loadfile.active.values)
    for act1 in actlist:
        if act1[0] == None:
            actlist.remove(act1)
    first = actlist[0]
    if first == ('No.', 'Blog Link', 'Username', 'Password') :
        print('Accounts input contains header, excluding header.')
        newlist = actlist[1:]
        return newlist
    elif ( not str(first[0]).isnumeric() ) and ( '.' not in str(first[1]) ):
        print('Accounts input contains header, excluding header.')
        newlist = actlist[1:]
        return newlist
    else:
        print('No header')
        return actlist


def loadcnt(xlsx):
    loadfile = openpyxl.load_workbook(xlsx)

    # *** check header *** #
    contentlist = list(loadfile.active.values)

    for cnt1 in contentlist:
        if cnt1[0] == None:
            contentlist.remove(cnt1)

    first = contentlist[0]
    if first == ('No.', 'Title', 'Content', 'Tags') :
        print('Content input contains header, excluding header.')
        newlist = contentlist[1:]
        return newlist
    elif ( not str(first[0]).isnumeric() ) and ( len(str(first[1])) < 15 ):
        print('Content input contains header, excluding header.\n')
        newlist = contentlist[1:]
        return newlist
    else:
        print('No header\n')
        return contentlist


def loaddel(xlsx):
    loadfile = openpyxl.load_workbook(xlsx)

    # *** check header *** #
    dellist = list(loadfile.active.values)

    for del1 in dellist:
        if del1[0] == None:
            dellist.remove(del1)

    first = dellist[0]
    if first == ('No.', 'Blog_url', 'Username', 'Password') :
        print('Content input contains header, excluding header.')
        newlist = dellist[1:]
        return newlist
    elif ( not str(first[0]).isnumeric() ) and ( len(str(first[1])) < 15 ):
        print('Content input contains header, excluding header.\n')
        newlist = dellist[1:]
        return newlist
    else:
        print('No header\n')
        return dellist


def navlogout():
    nlogout = 'https://nid.naver.com/nidlogin.logout'
    browser.get(nlogout)


def navlogin_1x(actdet):

    # *** naver url paths *** #
    nsite = 'https://section.blog.naver.com/BlogHome.nhn'
    nlogin = 'https://nid.naver.com/nidlogin.login'
    nblogpg = 'https://blog.naver.com/MyBlog.nhn'

    # *** parse input *** #
    actnum, bloglnk, user, pw = actdet
    print("Test Login to " + user + ' with password ' + pw)

    # *** get site *** #
    browser.get(nsite)
    browser.get(nlogin)
    try:
        myElem = WebDriverWait(browser, delay).until(EC.presence_of_element_located((By.ID, 'id')))
    except TimeoutException:
        print('Timeout. Unable to load page or internet is very slow.')

    # *** identify elements *** #
    el_user = browser.find_element_by_id('id')
    el_passw = browser.find_element_by_id('pw')
    el_logbtn = browser.find_element_by_xpath("//input[@type='submit'][@value='Sign in']")
    for character in user:
        el_user.send_keys(character)
        time.sleep(0.1)

    for character in pw:
        el_passw.send_keys(character)
        time.sleep(0.1)
    el_logbtn.click()
    time.sleep(1)

    # *** Captcha image : Check for presence *** #
    if chk_exist("//input[@type='submit'][@value='Sign in']"):
        print('\tLogin required captcha for ' + user + ' . Login manually to account, and confirm sign-in to new environment from mail section, to reduce captcha chances.')
    else:
        print('\tLogin Success for ' + user)

def navlogin(actdet):

    # *** naver url paths *** #
    nsite = 'https://section.blog.naver.com/BlogHome.nhn'
    nlogin = 'https://nid.naver.com/nidlogin.login'
    nblogpg = 'https://blog.naver.com/MyBlog.nhn'

    # *** parse input *** #
    actnum, bloglnk, user, pw = actdet
    print("Logging in to " + user + ' with password ' + pw)

    # *** get site *** #
    browser.get(nsite)
    browser.get(nlogin)
    try:
        WebDriverWait(browser, delay).until(EC.presence_of_element_located((By.ID, 'id')))
    except TimeoutException:
        print('Timeout. Unable to load page or internet is very slow.')

    # *** identify elements *** #
    el_user = browser.find_element_by_id('id')
    el_passw = browser.find_element_by_id('pw')
    el_logbtn = browser.find_element_by_xpath("//input[@type='submit'][@value='Sign in']")

    for character in user:
        el_user.send_keys(character)
        time.sleep(0.1)

    for character in pw:
        el_passw.send_keys(character)
        time.sleep(0.1)

    el_logbtn.click()
    time.sleep(1)

    # *** Captcha image : Check for presence
    # "//div[@class='captcha']"
    if chk_exist("//input[@type='submit'][@value='Sign in']"):
        updstat('Process: Start Blog Posting ... Captcha Required for Login.')
        while chk_exist("//input[@type='submit'][@value='Sign in']"):
            el_passw = browser.find_element_by_id('pw')
            el_passw.send_keys(pw)
            capt = input('Captcha required. Input manually and sign-in. Press enter to resume/retry.')
        if not chk_exist("//input[@type='submit'][@value='Sign in']"):
            print('Login Success.')
            updstat('Process: Start Blog Posting ... Login Success')
    else:
        print('Login Success.')
        updstat('Process: Start Blog Posting ... Login Success')

def retry_login():
    el_passw = browser.find_element_by_id('pw')
    el_passw.send_keys(pw)

def change_editor(actdet):
    # ***  Just for the username
    actnum, bloglnk, user, pw = actdet
    # ***  Change Editor Type
    nadmin = 'https://admin.blog.naver.com/' + user
    browser.get(nadmin)
    try:
        WebDriverWait(browser, delay).until(EC.presence_of_element_located((By.ID, 'default_editor_anchor')))
    except TimeoutException:
        print('Timeout. Unable to load page or internet is very slow.')


    el_edits = browser.find_element_by_id('default_editor_anchor')
    el_edits.click()

    browser.switch_to.frame(browser.find_element_by_xpath("//iframe[@id='papermain']"))
    browser.find_element_by_class_name('_selectSE2').click()
    btnsave = browser.find_element_by_id('submit_defaultEditor')
    btnsave.click()

def finpg():
    browser.get('https://thumbs.dreamstime.com/b/what-next-38031395.jpg')

def navpost(cntdet,actdet):
    delay = 5

    # *** naver url paths
    nsite = 'https://section.blog.naver.com/BlogHome.nhn'
    nlogin = 'https://nid.naver.com/nidlogin.login'
    nblogpg = 'https://blog.naver.com/MyBlog.nhn'

    # ***  parse input
    actnum, bloglnk, user, pw = actdet
    cntnum, intitle, incontent, intags = cntdet

    time.sleep(1)
    # *** write post
    nwrite = 'https://blog.naver.com/' + user + '/postwrite'
    browser.get(nwrite)
    try:
        alert = browser.switch_to.alert
        alert.accept()
    except:
        randvar = 'None'
    try:
        WebDriverWait(browser, delay).until(EC.presence_of_element_located((By.ID, 'subject')))
    except TimeoutException:
        print('Timeout. Unable to load page or internet is very slow.')

    # ***  add title
    el_title = browser.find_element_by_xpath("//input[@type='text'][@name='post.title']")
    el_title.clear()
    el_title.send_keys(intitle)

    # ***  add text
    switch_text = browser.find_element_by_class_name('se2_to_text')
    switch_text.click()
    alert = browser.switch_to.alert
    alert.accept()

    el_text = browser.find_element_by_class_name('se2_input_text')
    el_text.clear()
    el_text.send_keys(incontent)

    # ***  add tag
    el_tags = browser.find_element_by_id('tagList')
    el_tags.clear()
    el_tags.send_keys(intags)

    # ***  submit
    el_submit = browser.find_element_by_xpath("//img[@src='https://blogimgs.pstatic.net/nblog/mylog/post/btn_confirm02_2.gif']")
    el_submit.click()

    try:
        WebDriverWait(browser, delay).until(EC.presence_of_element_located((By.ID, 'printPost1')))
    except:
        if chk_exist("//input[@name='captchaKey']"):
            while chk_exist("//input[@name='captchaKey']"):
                capt = input('Captcha required for post. Input manually and press enter to retry submit post.')
            if not chk_exist("//input[@name='captchaKey']"):
                print('Captcha Success. Blog Posted.')
        else:
            print('Blog post fail reason unknown. Need to check with developer.')
    try:
        alert = browser.switch_to.alert
        alert.accept()
    except:
        randVar = None

    # *** hyperlink
    rgxlnk = re.compile('og:url.*\"(.*)\"')
    lnk = rgxlnk.search(browser.page_source).groups()[0]
    time.sleep(1)

    # *** get output
    browser.get(lnk)
    write_lnk(lnk)
    logme(lnk)
    print('Blog post done: ' + str(lnk))

def navdel(deldet):
    delay = 5

    # ***  parse input
    actnum, dellnk, user, pw = deldet

    try:
        browser.get(dellnk)
        time.sleep(1)

        try:
            alert = browser.switch_to.alert
            alert.accept()
        except:
            test = 'None'

        browser.switch_to.frame(browser.find_element_by_xpath("//iframe[@id='mainFrame']"))
        browser.find_element_by_class_name('_deletePost').click()

        try:
            alert = browser.switch_to.alert
            alert.accept()
        except:
            test = 'None'

        print('Done deleting/Already deleted. ' + str(dellnk))
        time.sleep(1)
    except:
        print('Failed to delete, already deleted or unable to load page.' + str(del1))

# ======================================================================================================== #
#                                                Tkinter Functions                                         #
# ======================================================================================================== #

def fileDialog_e1():
    filename = filedialog.askopenfilename(initialdir="/", title="Select A File", filetype=
    (("excel files", "*.xlsx"), ("all files", "*.*")))
    e1.delete(0,END)
    e1.insert(0,filename)

def fileDialog_e2():
    filename = filedialog.askopenfilename(initialdir="/", title="Select A File", filetype=
    (("excel files", "*.xlsx"), ("all files", "*.*")))
    e2.delete(0,END)
    e2.insert(0,filename)

def fileDialog_e3():
    filename = filedialog.askopenfilename(initialdir="/", title="Select A File", filetype=
    (("excel files", "*.xlsx"), ("all files", "*.*")))
    e3.delete(0,END)
    e3.insert(0,filename)

def get_content():
    cnt = cntfile.get()
    return cnt

def get_acnt():
    acnt = acntfile.get()
    return acnt

def get_del():
    delb = delfile.get()
    return delb

def check_inputs():
    updstat('Process: Checking Inputs ...')
    try:
        cntlist = loadact(cntfile.get())
        actlist = loadcnt(acntfile.get())
        updstat('Process: Checking Inputs ...   All inputs exist')
        time.sleep(1)
        updstat('Process: Checking Inputs ...   Test login once')
        for act in actlist:
            navlogin_1x(act)
        upd = 'Process: Checking Inputs Done.   Contents = ' + str(len(cntlist)) + '  Accounts = ' + str(len(actlist))
        updstat(upd)
    except:
        result = 'Some inputs do not exist. Recheck input'
        upd = 'Process: Checking Inputs ...   ' + result
        updstat(upd)
    finpg()

def thread_check_inputs():
    threading.Thread(target=check_inputs).start()

def exec_start():
    updstat('Process: Start Blog Posting ...')
    cntlist = []
    actlist = []
    try:
        cntlist = loadact(cntfile.get())
        actlist = loadcnt(acntfile.get())
    except:
        result = 'Some inputs do not exist. Recheck input'
        upd = 'Process: Checking Inputs ...   ' + result
        updstat(upd)

    actot = len(actlist)
    cntot = len(cntlist)

    actc = 0
    cntc = 0

    def updp():
        upd = 'Process: Posting Blogs...   Current Status:  Accounts( ' + str(actc) + '/' + str(actot) + ' )   Contents( ' + str(cntc) + '/' + str(cntot) + ' )'
        status.configure(text=upd)

    updp()
    for act in actlist:
        updp()
        navlogin(act)
        if chk_exist("//input[@type='submit'][@value='Sign in']"):
            print('Login Failed. Retry.')
            navlogin(act)
        change_editor(act)
        for cnt in cntlist:
            try:
                navpost(cnt, act)
            except:
                print("Blog Post Failed. Checking source of issue.")
                if chk_exist("//input[@type='submit'][@value='Sign in']"):
                    print("Forced Logout. Retry Login.")
                    navlogin(act)
                    navpost(cnt, act)
                else:
                    print("Issue can't be determined. Retrying to Post Blog anyway.")
                    navpost(cnt, act)
                continue
            cntc += 1
            updp()
        navlogout()
        actc += 1
        updp()
    #upd = 'Process: Blog Posting Done. Check output from outputs.txt file inside same directory.'
    upd = 'Process: Blog Posting Done.   Accounts( ' + str(actc) + '/' + str(actot) + ' )   Contents( ' + str(cntc) + '/' + str(cntot) + ' ) .  Check outputs.txt file'
    updstat(upd)
    finpg()

def thread_start():
    threading.Thread(target=exec_start).start()


def exec_del():
    dellist = []
    delc = 0

    updstat('Process: Deleting Blog Posts ...')
    try:
        dellist = loaddel(delfile.get())
    except:
        result = 'Some inputs do not exist. Recheck input'
        upd = 'Process: Deleting Blog Posts ...   ' + result
        updstat(upd)
    deltot = len(dellist)

    def updd():
        upd = 'Process: Deleting Posts...   Delete Status: ( ' + str(delc) + '/' + str(deltot) + ' )'
        status.configure(text=upd)

    navlogin(dellist[0])
    current = dellist[0][2]
    updd()
    for del1 in dellist:
        if current == del1[2]:
            navdel(del1)
        else:
            navlogin(del1)
            navdel(del1)
        updd()
        delc += 1
        current = del1[2]
    upd = 'Process: Deleting Posts Done. Delete Status: ( ' + str(delc) + '/' + str(deltot) + ' )'
    updstat(upd)
    finpg()
def thread_del():
    threading.Thread(target=exec_del).start()
# ======================================================================================================== #
#                                                Captcha Input                                             #
# ======================================================================================================== #
'''
import keyboard
class MyDialog:

    def __init__(self, parent):
        top = self.top = tk.Toplevel(parent)
        self.myLabel = tk.Label(top, text='Input captcha below:')
        self.myLabel.pack()
        self.myEntryBox = tk.Entry(top)
        self.myEntryBox.pack()
        self.mySubmitButton = tk.Button(top, text='Submit', command=self.send)
        self.mySubmitButton.pack()

    def send(self):
        self.username = self.myEntryBox.get()
        self.top.destroy()

def onClick():
    inputDialog = MyDialog(root)
    root.wait_window(inputDialog.top)
    print('Username: ', inputDialog.username)

def cmd2():
    keyboard.press_and_release('enter')

def thread_cmd2():
    threading.Thread(target=cmd2).start()
'''
# ======================================================================================================== #
#                                                Tkinter GUI                                               #
# ======================================================================================================== #


# *** Input Files Section
Label(master, text="Input Content", fg="green").grid(row=0, sticky=E)
Label(master, text="Input Accounts", fg="green").grid(row=1, sticky=E)
Label(master, text="Delete Blogs", fg="green").grid(row=2, sticky=E)

cntfile = StringVar()
acntfile = StringVar()
delfile = StringVar()

e1 = Entry(master, textvariable=cntfile)
e2 = Entry(master, textvariable=acntfile)
e3 = Entry(master, textvariable=delfile)
e1.grid(row=0, column=1, sticky=W+E)
e2.grid(row=1, column=1, sticky=W+E)
e3.grid(row=2, column=1, sticky=W+E)

brws1 = Button(master, text='browse file', fg="blue", command=fileDialog_e1)
brws2 = Button(master, text='browse file', fg="blue", command=fileDialog_e2)
brws3 = Button(master, text='browse file', fg="blue", command=fileDialog_e3)
brws1.grid(row=0, column=3, sticky=W+E)
brws2.grid(row=1, column=3, sticky=W+E)
brws3.grid(row=2, column=3, sticky=W+E)

# ***  Execute Buttons Section
frame = Frame(master)
frame.grid(row=3, columnspan=3)

xb1 = Button(frame, text='Check Inputs', command=thread_check_inputs, height=1, width=15, fg="green")
xb2 = Button(frame, text='Start', command=thread_start, height=1, width=7, fg="green")
xb3 = Button(frame, text='Stop', command=master.quit, height=1, width=7, fg="red")

xbquit = Button(frame, text='Quit', command=master.quit, height=1, width=7)
xbd = Button(frame, text='Delete Posts(Careful)', command=thread_del, height=1, width=17, fg="orange")
#xbret = Button(frame, text='Retry/Resume', command=thread_cmd2, height=1, width=7, fg="green")

#xbquit = Button(frame, text='Quit', command=master.quit, height=1, width=7)

xb1.grid(row=0, column=1, padx=1, pady=1, sticky=W+E)
xb2.grid(row=0, column=2, padx=1, pady=1, sticky=W+E)
#xb3.grid(row=0, column=3, padx=1, pady=1, sticky=W+E)
xbquit.grid(row=0, column=4, padx=1, pady=1, sticky=W+E)
xbd.grid(row=0, column=5, padx=47, pady=1, sticky=W+E)
#xbret.grid(row=0, column=6, padx=1, pady=1, sticky=W+E)

# *** Status Bar
status = Label(master, text="Process : Initialized ", bd=1, relief=SUNKEN, anchor=W)
status.grid(row=4, columnspan=4, sticky=W+E)

# *** temp add files
#e1.insert(0,'C:/Users/ersantiago/Desktop/naver/files/input_content_3.xlsx')
#e2.insert(0,'C:/Users/ersantiago/Desktop/naver/files/accounts_1.xlsx')
#e3.insert(0,'C:/Users/ersantiago/Desktop/naver/files/delete_list.xlsx')
master.mainloop()

'''
browser = webdriver.Chrome('driver\chromedriver')

# VARIABLES & PATHS
file_cnt = 'C:/Users/ersantiago/Desktop/naver/files/input_content_3.xlsx'
file_act = 'C:/Users/ersantiago/Desktop/naver/files/accounts_1.xlsx'
file_del = 'C:/Users/ersantiago/Desktop/naver/files/delete_list.xlsx'

# Load Inputs
cntlist = loadcnt(file_cnt)
actlist = loadact(file_act)
dellist = loaddel(file_del)

cnt1 = cntlist[0]
act1 = actlist[0]
del1 = dellist[0]

navlogin(act1)
navpost(cnt1,act1)


# Post Captcha
browser.find_element_by_id('captchalayeredframe')
browser.switch_to.frame(browser.find_element_by_xpath("//iframe[@id='captchalayeredframe']"))
captchabox = browser.find_element_by_id('captchaValue')
captchaimg = browser.find_element_by_xpath("//img[@id='captchaImage']").get_attribute('src')
try:
    alert = browser.switch_to.alert
    alert.accept()
except:
    randVar = None

'''